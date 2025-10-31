# validate_exact_cells.py
# -*- coding: utf-8 -*-

from __future__ import annotations
from pathlib import Path
import sys, math
import os
from typing import Any, Dict, List, Tuple, Optional, Union
from openpyxl import load_workbook

from beitrag_und_verlaufswerte import (
    TarifInput,
    beitragsberechnung,
    verlaufswerte,
)

THIS_DIR   = Path(__file__).resolve().parent        # …/Arno/output
INPUT_DIR  = THIS_DIR.parent / "input"              # …/Arno/input
os.chdir(str(THIS_DIR))                            # CWD = output

EXCEL_PATH_DEFAULT = INPUT_DIR / "Tarifrechner_KLV.xlsm"
SHEET_NAME = "Kalkulation"

# ---------- Parsing-Helfer ----------
def _txt(v: Any) -> str:
    return "" if v is None else str(v).strip()

def num_de(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "-"):
        return 0.0
    s = (s.replace("€", "").replace("%", "").replace(" ", "").replace("\xa0", ""))
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0

def pct2dec(raw: Any, val: float) -> float:
    return val / 100.0 if isinstance(raw, str) and "%" in raw else val

# ---------- Feste Zelladressen (exakt wie im Screenshot) ----------

# Vertragsdaten (Werte in Spalte B, ab Zeile 4)
CELL_X   = "B4"
CELL_SEX = "B5"
CELL_N   = "B6"
CELL_T   = "B7"
CELL_VS  = "B8"
CELL_ZW  = "B9"

# Tarifdaten (Werte in Spalte E, ab Zeile 4)
CELL_ZINS   = "E4"
CELL_TAFEL  = "E5"
CELL_ALPHA  = "E6"
CELL_BETA1  = "E7"
CELL_GAMMA1 = "E8"
CELL_GAMMA2 = "E9"
CELL_GAMMA3 = "E10"
CELL_K      = "E11"
CELL_RATZU  = "E12"

# Grenzen (Werte in Spalte H, ab Zeile 4)
CELL_MINALTERFLEX = "H4"
CELL_MINRLZFLEX   = "H5"

# Beitragsberechnung (rechte Seite; Werte in Spalte K)
# Block beginnt in Zeile 5 und hat eine Leerzeile vor Pxt:
#   Bxt -> K5
#   BJB -> K6
#   BZB -> K7
#   Pxt -> K9
BEITRAG_K = {"Bxt": "K5", "BJB": "K6", "BZB": "K7", "Pxt": "K9"}

# Verlaufswerte (Kopfzeile Zeile 15, Daten ab Zeile 16)
HEADER_ROW = 15
DATA_START = 16
COL = {
    "k": "A",
    "Axn": "B",
    "axn": "C",
    "axt": "D",
    "kVx_bpfl": "E",
    "kDRx_bpfl": "F",
    "kVx_bfr": "G",
    "kVx_MRV": "H",
    "flex_phase": "I",  # "flex. Phase"
    "StoAb": "J",
    "RKW": "K",
    "VS_bfr": "L",
}

# ---------- Excel lesen ----------
def read_inputs(path: Path) -> TarifInput:
    wb = load_workbook(path, data_only=True, read_only=True, keep_vba=True)
    ws = wb[SHEET_NAME]

    x   = int(num_de(ws[CELL_X].value))
    sex = _txt(ws[CELL_SEX].value) or "M"
    n   = int(num_de(ws[CELL_N].value))
    t   = int(num_de(ws[CELL_T].value))
    VS  = num_de(ws[CELL_VS].value)
    zw  = int(num_de(ws[CELL_ZW].value))

    z_raw = ws[CELL_ZINS].value;   zins   = pct2dec(z_raw,   num_de(z_raw))
    tafel = _txt(ws[CELL_TAFEL].value)
    a_raw = ws[CELL_ALPHA].value;  alpha  = pct2dec(a_raw,  num_de(a_raw))
    b_raw = ws[CELL_BETA1].value;  beta1  = pct2dec(b_raw,  num_de(b_raw))
    g1r   = ws[CELL_GAMMA1].value; gamma1 = pct2dec(g1r,    num_de(g1r))
    g2r   = ws[CELL_GAMMA2].value; gamma2 = pct2dec(g2r,    num_de(g2r))
    g3r   = ws[CELL_GAMMA3].value; gamma3 = pct2dec(g3r,    num_de(g3r))
    k     = num_de(ws[CELL_K].value)
    r_raw = ws[CELL_RATZU].value;  ratzu  = pct2dec(r_raw,  num_de(r_raw))

    MinAlterFlex = int(num_de(ws[CELL_MINALTERFLEX].value))
    MinRLZFlex   = int(num_de(ws[CELL_MINRLZFLEX].value))

    wb.close()
    return TarifInput(
        x=x, sex=sex, n=n, t=t, VS=VS, zw=zw,
        zins=zins, tafel=tafel, alpha=alpha, beta1=beta1,
        gamma1=gamma1, gamma2=gamma2, gamma3=gamma3,
        k=k, ratzu=ratzu,
        MinAlterFlex=MinAlterFlex, MinRLZFlex=MinRLZFlex
    )

def read_expected_beitrag(path: Path) -> Dict[str, float]:
    wb = load_workbook(path, data_only=True, read_only=True, keep_vba=True)
    ws = wb[SHEET_NAME]
    out = {k: num_de(ws[cell].value) for k, cell in BEITRAG_K.items()}
    wb.close()
    return out

def read_expected_verlauf(path: Path) -> List[Dict[str, float]]:
    wb = load_workbook(path, data_only=True, read_only=True, keep_vba=True)
    ws = wb[SHEET_NAME]
    rows: List[Dict[str, float]] = []
    r = DATA_START
    while True:
        k_cell = ws[f"{COL['k']}{r}"].value
        if k_cell in (None, ""):
            break
        row = {
            "k": int(num_de(k_cell)),
            "Axn":       num_de(ws[f"{COL['Axn']}{r}"].value),
            "axn":       num_de(ws[f"{COL['axn']}{r}"].value),
            "axt":       num_de(ws[f"{COL['axt']}{r}"].value),
            "kVx_bpfl":  num_de(ws[f"{COL['kVx_bpfl']}{r}"].value),
            "kDRx_bpfl": num_de(ws[f"{COL['kDRx_bpfl']}{r}"].value),
            "kVx_bfr":   num_de(ws[f"{COL['kVx_bfr']}{r}"].value),
            "kVx_MRV":   num_de(ws[f"{COL['kVx_MRV']}{r}"].value),
            "flex_phase": int(num_de(ws[f"{COL['flex_phase']}{r}"].value)),
            "StoAb":     num_de(ws[f"{COL['StoAb']}{r}"].value),
            "RKW":       num_de(ws[f"{COL['RKW']}{r}"].value),
            "VS_bfr":    num_de(ws[f"{COL['VS_bfr']}{r}"].value),
        }
        rows.append(row)
        r += 1
    wb.close()
    return rows

# ---------- Vergleich ----------
def close(a: float, b: float, rel=1e-9, abs_tol=1e-10) -> bool:
    return math.isclose(a, b, rel_tol=rel, abs_tol=abs_tol)

def cmp(name: str, calc: float, exp: float, money=False,
        rel=1e-9, abs_tol=None) -> Tuple[bool, str]:
    if abs_tol is None:
        abs_tol = 0.01 if money else 1e-10
    ok = close(calc, exp, rel, abs_tol)
    suf = " €" if money else ""
    if ok:
        return True, f"[OK] {name}: {calc:.10f}{suf}"
    return False, f"[DIFF] {name}: calc={calc:.10f}{suf} excel={exp:.10f}{suf} Δ={calc-exp:.10f}{suf}"

# ---------- Main ----------
def main(xlsm_path: Optional[Union[str, Path]] = None) -> int:
    # Kein Zugriff auf sys.argv hier!
    path = Path(EXCEL_PATH_DEFAULT) if xlsm_path is None else Path(xlsm_path)
    if not path.exists():
        sys.stderr.write(f"Datei nicht gefunden: {path}\n")
        sys.stderr.write("Aufruf: python vergleich.py [Pfad/zur/Tarifrechner_KLV.xlsm]\n")
        return 1

    ti = read_inputs(path)
    beitr_calc = beitragsberechnung(ti)
    verl_calc  = verlaufswerte(ti)

    beitr_exp = read_expected_beitrag(path)
    verl_exp  = read_expected_verlauf(path)

    print("=== Beitragsberechnung (K5,K6,K7,K9) ===")
    ok_all = True
    for key, money in (("Bxt", False), ("BJB", True), ("BZB", True), ("Pxt", False)):
        ok, msg = cmp(key, beitr_calc[key], beitr_exp.get(key, float("nan")), money)
        print(msg)
        ok_all &= ok

    print("\n=== Verlaufswerte (Header Zeile 15, Daten ab Zeile 16) ===")
    n = min(len(verl_calc), len(verl_exp))
    cols = [
        ("Axn", False), ("axn", False), ("axt", False), ("kVx_bpfl", False),
        ("kDRx_bpfl", True), ("kVx_bfr", False), ("kVx_MRV", True),
        ("StoAb", True), ("RKW", True), ("VS_bfr", True),
    ]
    for i in range(n):
        c, e = verl_calc[i], verl_exp[i]
        prefix = f"k={c['k']}: "
        if c["k"] != e["k"]:
            print(prefix + f"[DIFF] k: calc={c['k']} excel={e['k']}")
            ok_all = False
        if int(c["flex_phase"]) == int(e["flex_phase"]):
            print(prefix + f"[OK] flex_phase={c['flex_phase']}")
        else:
            print(prefix + f"[DIFF] flex_phase: calc={c['flex_phase']} excel={e['flex_phase']}")
            ok_all = False
        for col, money in cols:
            ok, msg = cmp(prefix + col, float(c[col]), float(e[col]), money)
            print(msg)
            ok_all &= ok

    print("\n=== Gesamtergebnis ===")
    print("PASS" if ok_all else "FAIL")
    return 0 if ok_all else 2

if __name__ == "__main__":
    sys.exit(main())
