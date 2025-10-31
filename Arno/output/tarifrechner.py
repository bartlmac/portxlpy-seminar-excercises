# main.py
# -*- coding: utf-8 -*-

import sys
import os
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook 

# Dein Modul aus dem vorherigen Schritt:
from beitrag_und_verlaufswerte import (
    TarifInput,
    beitragsberechnung,
    verlaufswerte,
)

THIS_DIR   = Path(__file__).resolve().parent        # …/Arno/output
INPUT_DIR  = THIS_DIR.parent / "input"              # …/Arno/input
os.chdir(str(THIS_DIR))                            # CWD = output

EXCEL_PATH_DEFAULT = INPUT_DIR / "Tarifrechner_KLV.xlsm"
SHEET_NAME = "Kalkulation"


# ---------------------- Excel-Utilities ----------------------

def _norm_text(v: Any) -> str:
    return str(v).strip() if v is not None else ""

def _parse_number_de(v: Any) -> float:
    """
    Robust gegen deutsche Formatierung (Punkt = Tausender, Komma = Dezimal),
    sowie gegen Suffixe wie '€' und '%'.
    """
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)

    s = str(v).strip()
    if s == "":
        return 0.0

    # Entferne Währungs-/Prozent-/Leerzeichen
    s = (
        s.replace("€", "")
         .replace("%", "")
         .replace(" ", "")
         .replace("\xa0", "")   # NBSP
    )
    # Tausenderpunkte raus, Dezimalkomma -> Punkt
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        # Falls doch nicht numerisch (z. B. 'DAV1994_T'), zurück 0.0
        return 0.0

def find_right_value(ws, label: str) -> Optional[Any]:
    """
    Sucht im gesamten Blatt nach einer Zelle mit exakt `label`
    (nach Trim), und liefert den Wert der Zelle rechts daneben (gleiche Zeile, +1 Spalte).
    """
    target = label.strip()
    for row in ws.iter_rows():
        for cell in row:
            if _norm_text(cell.value) == target:
                right = ws.cell(row=cell.row, column=cell.column + 1)
                return right.value
    return None

def get_string(ws, label: str) -> str:
    v = find_right_value(ws, label)
    return "" if v is None else str(v).strip()

def get_number(ws, label: str) -> float:
    v = find_right_value(ws, label)
    return _parse_number_de(v)


# ---------------------- Hauptlogik ----------------------

def read_inputs_from_excel(xlsm_path: str | Path) -> TarifInput:
    wb = load_workbook(filename=xlsm_path, data_only=True, read_only=True, keep_vba=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Tabellenblatt '{SHEET_NAME}' nicht gefunden.")

    ws = wb[SHEET_NAME]

    # Vertragsdaten
    x      = int(get_number(ws, "x"))
    sex    = get_string(ws, "Sex") or "M"
    n      = int(get_number(ws, "n"))
    t      = int(get_number(ws, "t"))
    VS     = get_number(ws, "VS")
    zw     = int(get_number(ws, "zw"))

    # Tarifdaten
    zins   = get_number(ws, "Zins") / 100.0 if "%" in str(find_right_value(ws, "Zins")) else get_number(ws, "Zins")
    tafel  = get_string(ws, "Tafel")
    alpha  = get_number(ws, "alpha")  / 100.0 if "%" in str(find_right_value(ws, "alpha"))  else get_number(ws, "alpha")
    beta1  = get_number(ws, "beta1")  / 100.0 if "%" in str(find_right_value(ws, "beta1"))  else get_number(ws, "beta1")
    gamma1 = get_number(ws, "gamma1") / 100.0 if "%" in str(find_right_value(ws, "gamma1")) else get_number(ws, "gamma1")
    gamma2 = get_number(ws, "gamma2") / 100.0 if "%" in str(find_right_value(ws, "gamma2")) else get_number(ws, "gamma2")
    gamma3 = get_number(ws, "gamma3") / 100.0 if "%" in str(find_right_value(ws, "gamma3")) else get_number(ws, "gamma3")
    k      = get_number(ws, "k")
    ratzu  = get_number(ws, "ratzu") / 100.0 if "%" in str(find_right_value(ws, "ratzu")) else get_number(ws, "ratzu")

    # Grenzen
    MinAlterFlex = int(get_number(ws, "MinAlterFlex"))
    MinRLZFlex   = int(get_number(ws, "MinRLZFlex"))

    wb.close()

    return TarifInput(
        x=x, sex=sex, n=n, t=t, VS=VS, zw=zw,
        zins=zins, tafel=tafel, alpha=alpha, beta1=beta1,
        gamma1=gamma1, gamma2=gamma2, gamma3=gamma3,
        k=k, ratzu=ratzu,
        MinAlterFlex=MinAlterFlex, MinRLZFlex=MinRLZFlex
    )


def format_eur(x: float) -> str:
    return f"{x:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")

def format_num(x: float, digits: int = 6) -> str:
    return f"{x:.{digits}f}".replace(".", ",")

def print_results(inp: TarifInput):
    # Beitragsberechnung
    beitr = beitragsberechnung(inp)
    print("\n=== Beitragsberechnung ===")
    print(f"Bxt: {format_num(beitr['Bxt'], 8)}")
    print(f"BJB: {format_eur(beitr['BJB'])}")
    print(f"BZB: {format_eur(beitr['BZB'])}")
    print(f"Pxt: {format_num(beitr['Pxt'], 8)}")

    # Verlaufswerte
    print("\n=== Verlaufswerte ===")
    header = [
        "k", "Axn", "axn", "axt", "kVx_bpfl", "kDRx_bpfl",
        "kVx_bfr", "kVx_MRV", "flex.Phase", "StoAb", "RKW", "VS_bfr"
    ]
    print("; ".join(header))

    rows = verlaufswerte(inp)  # bis max(n,t)
    for r in rows:
        line = [
            str(r["k"]),
            format_num(r["Axn"], 6),
            format_num(r["axn"], 6),
            format_num(r["axt"], 6),
            format_num(r["kVx_bpfl"], 6),
            format_eur(r["kDRx_bpfl"]),
            format_num(r["kVx_bfr"], 6),
            format_eur(r["kVx_MRV"]),
            str(r["flex_phase"]),
            format_eur(r["StoAb"]),
            format_eur(r["RKW"]),
            format_eur(r["VS_bfr"]) if r["VS_bfr"] >= 1000 else format_num(r["VS_bfr"], 2),
        ]
        print("; ".join(line))


def main():
    path = Path(EXCEL_PATH_DEFAULT)
    if len(sys.argv) > 1:
        path = Path(sys.argv[1])

    if not path.exists():
        sys.stderr.write(f"Datei nicht gefunden: {path}\n")
        sys.stderr.write("Aufruf: python main.py [Pfad/zur/Tarifrechner_KLV.xlsm]\n")
        sys.exit(1)

    inp = read_inputs_from_excel(path)
    print_results(inp)


if __name__ == "__main__":
    main()
